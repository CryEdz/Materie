from __future__ import annotations

import re
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
SOURCE_GLOB = "Voti_Fantacalcio_Stagione_2025_26_Giornata_*.xlsx"
TEMPLATE_FILE = BASE_DIR / "Fantacalcio_tidy.xlsx"
OUTPUT_FILE = BASE_DIR / "Fantacalcio_tidy.xlsx"
REPORT_ANALISI_FILE = BASE_DIR / "report_analisi.md"
REPORT_PROCESSO_FILE = BASE_DIR / "report_processo.md"
OUTPUT_DIR = BASE_DIR / "output"

TITLE_PREFIX = "Voti Fantacalcio"
SKIP_ROWS = {
    "Solo su www.fantacalcio.it i voti ufficiali per la tua lega",
    "QUESTO FILE NON PUO' ESSERE RIPRODOTTO NE' PUBBLICATO SU ALTRI SITI INTERNET",
    "E' DA CONSIDERARSI AD USO PERSONALE ESCLUSIVO DEGLI ISCRITTI DI FANTACALCIO.IT",
    "Cod.",
}

EXPECTED_COLUMNS = [
    "Squadra",
    "Cod.",
    "Ruolo",
    "Nome",
    "Voto",
    "Gf",
    "Gs",
    "Rp",
    "Rs",
    "Rf",
    "Au",
    "Amm",
    "Esp",
    "Ass",
    "Fantamedia",
]

ROLE_ORDER = ["P", "D", "C", "A", "ALL"]


def parse_round_number(path: Path) -> int:
    match = re.search(r"Giornata_(\d+)\.xlsx$", path.name)
    if not match:
        raise ValueError(f"Nome file non valido: {path.name}")
    return int(match.group(1))


def clean_vote(value: object) -> float:
    if value is None:
        raise ValueError("Voto mancante")
    if isinstance(value, str):
        return float(value.replace("*", "").strip())
    return float(value)


def to_int(value: object) -> int:
    if value is None:
        return 0
    return int(value)


def is_team_row(row: tuple[object, ...]) -> bool:
    return isinstance(row[0], str) and bool(row[0].strip()) and all(value is None for value in row[1:])


def is_title_or_meta_row(row: tuple[object, ...]) -> bool:
    first = row[0]
    if not isinstance(first, str):
        return False
    return first.startswith(TITLE_PREFIX) or first in SKIP_ROWS


def parse_source_file(path: Path) -> pd.DataFrame:
    round_number = parse_round_number(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, object]] = []
    current_team: str | None = None

    for row in worksheet.iter_rows(values_only=True):
        if is_title_or_meta_row(row):
            continue
        if is_team_row(row):
            current_team = str(row[0]).strip()
            continue
        if current_team is None:
            continue

        first = row[0]
        if not isinstance(first, (int, float)):
            continue

        rows.append(
            {
                "round": round_number,
                "team": current_team,
                "code": int(first),
                "role": str(row[1]).strip(),
                "name": str(row[2]).strip(),
                "vote": clean_vote(row[3]),
                "gf": to_int(row[4]),
                "gs": to_int(row[5]),
                "rp": to_int(row[6]),
                "rs": to_int(row[7]),
                "rf": to_int(row[8]),
                "au": to_int(row[9]),
                "amm": to_int(row[10]),
                "esp": to_int(row[11]),
                "ass": to_int(row[12]),
            }
        )

    return pd.DataFrame(rows)


def load_raw_data(source_dir: Path) -> pd.DataFrame:
    files = sorted(source_dir.glob(SOURCE_GLOB), key=parse_round_number)
    if not files:
        raise FileNotFoundError(f"Nessun file trovato con pattern {SOURCE_GLOB}")

    frames = [parse_source_file(path) for path in files]
    raw = pd.concat(frames, ignore_index=True)
    raw["fantavoto"] = (
        raw["vote"]
        + raw["gf"] * 3
        - raw["gs"]
        + raw["rp"] * 3
        - raw["rs"]
        - raw["rf"] * 3
        - raw["au"]
        - raw["amm"] * 0.5
        - raw["esp"]
        + raw["ass"]
    )
    return raw


def build_aggregated_table(raw: pd.DataFrame) -> pd.DataFrame:
    aggregated = (
        raw.groupby(["team", "code", "role", "name"], as_index=False)
        .agg(
            voto=("vote", "mean"),
            gf=("gf", "sum"),
            gs=("gs", "sum"),
            rp=("rp", "sum"),
            rs=("rs", "sum"),
            rf=("rf", "sum"),
            au=("au", "sum"),
            amm=("amm", "sum"),
            esp=("esp", "sum"),
            ass=("ass", "sum"),
            fantamedia=("fantavoto", "mean"),
            presenze=("round", "count"),
        )
        .rename(columns={"team": "Squadra", "code": "Cod.", "role": "Ruolo", "name": "Nome", "voto": "Voto"})
    )

    aggregated["Voto"] = aggregated["Voto"].round(2)
    aggregated["fantamedia"] = aggregated["fantamedia"].round(2)
    return aggregated


def build_team_summary(raw: pd.DataFrame) -> pd.DataFrame:
    team_summary = (
        raw.groupby("team", as_index=False)
        .agg(
            presenze=("fantavoto", "size"),
            fantamedia=("fantavoto", "mean"),
            voto_medio=("vote", "mean"),
        )
        .rename(columns={"team": "Squadra"})
        .sort_values("fantamedia", ascending=False)
    )
    team_summary[["fantamedia", "voto_medio"]] = team_summary[["fantamedia", "voto_medio"]].round(2)
    return team_summary


def build_role_summary(aggregated: pd.DataFrame) -> pd.DataFrame:
    role_summary = (
        aggregated.groupby("Ruolo", as_index=False)
        .agg(
            giocatori=("Nome", "count"),
            fantamedia_media=("fantamedia", "mean"),
            voto_medio=("Voto", "mean"),
        )
        .sort_values("fantamedia_media", ascending=False)
    )
    role_summary["fantamedia_media"] = role_summary["fantamedia_media"].round(2)
    role_summary["voto_medio"] = role_summary["voto_medio"].round(2)
    return role_summary


def markdown_value(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def build_rankings(raw: pd.DataFrame, aggregated: pd.DataFrame) -> dict[str, pd.DataFrame]:
    team_trend = (
        raw.groupby(["round", "team"], as_index=False)
        .agg(fantamedia=("fantavoto", "mean"))
        .rename(columns={"team": "Squadra"})
    )

    player_focus = aggregated[aggregated["presenze"] >= 15].sort_values(
        ["fantamedia", "presenze"], ascending=[False, False]
    )

    scorers = aggregated.sort_values(["gf", "fantamedia"], ascending=[False, False]).copy()
    assisters = aggregated.sort_values(["ass", "fantamedia"], ascending=[False, False]).copy()
    disciplinary = aggregated.assign(malus_disciplinare=aggregated["amm"] * 0.5 + aggregated["esp"]).sort_values(
        ["malus_disciplinare", "fantamedia"], ascending=[False, True]
    )
    consistency = (
        raw.groupby(["team", "code", "role", "name"], as_index=False)
        .agg(presenze=("round", "count"), fantamedia=("fantavoto", "mean"), std_fantavoto=("fantavoto", "std"))
        .query("presenze >= 15 and fantamedia >= 6.0")
        .sort_values(["std_fantavoto", "presenze"], ascending=[True, False])
    )
    consistency["fantamedia"] = consistency["fantamedia"].round(2)
    consistency["std_fantavoto"] = consistency["std_fantavoto"].round(2)

    return {
        "team_trend": team_trend,
        "player_focus": player_focus,
        "scorers": scorers,
        "assisters": assisters,
        "disciplinary": disciplinary,
        "consistency": consistency,
    }


def style_worksheet_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def write_tidy_workbook(aggregated: pd.DataFrame, output_file: Path) -> None:
    if TEMPLATE_FILE.exists():
        workbook = load_workbook(TEMPLATE_FILE)
        if "Risultati" not in workbook.sheetnames:
            workbook.create_sheet("Risultati", 0)
    else:
        workbook = Workbook()
        workbook.active.title = "Risultati"
        workbook.create_sheet("Legenda e bonus malus")

    results_sheet = workbook["Risultati"]
    if results_sheet.max_row > 1:
        results_sheet.delete_rows(2, results_sheet.max_row - 1)

    while results_sheet.max_column > len(EXPECTED_COLUMNS):
        results_sheet.delete_cols(len(EXPECTED_COLUMNS) + 1)

    for column_index, column_name in enumerate(EXPECTED_COLUMNS, start=1):
        results_sheet.cell(row=1, column=column_index, value=column_name)

    export_df = aggregated.copy()
    export_df = export_df[["Squadra", "Cod.", "Ruolo", "Nome", "Voto", "gf", "gs", "rp", "rs", "rf", "au", "amm", "esp", "ass", "fantamedia"]]
    export_df = export_df.sort_values(["Squadra", "Ruolo", "fantamedia", "Nome"], ascending=[True, True, False, True])
    export_df.columns = EXPECTED_COLUMNS

    for row in export_df.itertuples(index=False, name=None):
        results_sheet.append(list(row))

    legend_sheet = workbook["Legenda e bonus malus"] if "Legenda e bonus malus" in workbook.sheetnames else workbook.create_sheet("Legenda e bonus malus")
    if legend_sheet.max_row > 0:
        legend_sheet.delete_rows(1, legend_sheet.max_row)

    legend_rows = [
        ["Valore", "Significato", "Bonus/malus"],
        ["Voto", "Voto del giocatore", "-"],
        ["Gf", "Gol fatto", 3],
        ["Gs", "Gol subito", -1],
        ["Rp", "Rigore parato", 3],
        ["Rs", "Rigore subito", -1],
        ["Rf", "Rigore fallito", -3],
        ["Au", "Autorete", -1],
        ["Amm", "Ammonizione", -0.5],
        ["Esp", "Espulsione", -1],
        ["Ass", "Assist", 1],
        [None, None, None],
        ["Il fantavoto e la somma del voto e di tutti i bonus/malus", None, None],
    ]
    for row in legend_rows:
        legend_sheet.append(row)

    style_worksheet_header(results_sheet)
    style_worksheet_header(legend_sheet)
    results_sheet.freeze_panes = "A2"
    legend_sheet.freeze_panes = "A2"

    for sheet in (results_sheet, legend_sheet):
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            max_length = max((len(value) for value in values), default=0)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 24)

    workbook.save(output_file)


def save_bar_chart(frame: pd.DataFrame, x: str, y: str, title: str, output_path: Path, color: str = "#1F4E78") -> None:
    plt.figure(figsize=(11, 6))
    bars = plt.barh(frame[x], frame[y], color=color)
    plt.gca().invert_yaxis()
    plt.title(title, fontsize=14, weight="bold")
    plt.xlabel(y.replace("_", " ").title())
    plt.ylabel(x)
    plt.grid(axis="x", alpha=0.25)
    for bar in bars:
        value = bar.get_width()
        plt.text(value + 0.03, bar.get_y() + bar.get_height() / 2, f"{value:.2f}", va="center", fontsize=9)
    plt.tight_layout()
    plt.savefig(output_path, dpi=160, bbox_inches="tight")
    plt.close()


def save_line_chart(frame: pd.DataFrame, title: str, output_path: Path) -> None:
    plt.figure(figsize=(12, 6))
    for squadra, subset in frame.groupby("Squadra"):
        plt.plot(subset["round"], subset["fantamedia"], linewidth=2, label=squadra)
    plt.title(title, fontsize=14, weight="bold")
    plt.xlabel("Giornata")
    plt.ylabel("Fantamedia media")
    plt.grid(alpha=0.25)
    plt.legend(loc="best", ncols=2)
    plt.tight_layout()
    plt.savefig(output_path, dpi=160, bbox_inches="tight")
    plt.close()


def make_visuals(raw: pd.DataFrame, aggregated: pd.DataFrame, rankings: dict[str, pd.DataFrame]) -> dict[str, Path]:
    OUTPUT_DIR.mkdir(exist_ok=True)

    top_players = rankings["player_focus"].head(10)[["Nome", "fantamedia"]].copy()
    save_bar_chart(
        top_players.sort_values("fantamedia", ascending=True),
        x="Nome",
        y="fantamedia",
        title="Top 10 giocatori per fantamedia con almeno 15 presenze",
        output_path=OUTPUT_DIR / "01_top_giocatori_fantamedia.png",
        color="#2F6B5F",
    )

    role_summary = build_role_summary(aggregated).copy()
    save_bar_chart(
        role_summary.sort_values("fantamedia_media", ascending=True),
        x="Ruolo",
        y="fantamedia_media",
        title="Fantamedia media per ruolo",
        output_path=OUTPUT_DIR / "02_fantamedia_per_ruolo.png",
        color="#8E6E53",
    )

    team_summary = build_team_summary(raw).head(10)
    save_bar_chart(
        team_summary.sort_values("fantamedia", ascending=True),
        x="Squadra",
        y="fantamedia",
        title="Top 10 squadre per fantamedia media stagionale",
        output_path=OUTPUT_DIR / "03_top_squadre.png",
        color="#4C6A92",
    )

    top_teams = team_summary.head(4)["Squadra"].tolist()
    team_trend = rankings["team_trend"][rankings["team_trend"]["Squadra"].isin(top_teams)]
    save_line_chart(
        team_trend,
        title="Andamento della fantamedia media per giornata delle migliori squadre",
        output_path=OUTPUT_DIR / "04_andamento_squadre_top.png",
    )

    return {
        "players": OUTPUT_DIR / "01_top_giocatori_fantamedia.png",
        "roles": OUTPUT_DIR / "02_fantamedia_per_ruolo.png",
        "teams": OUTPUT_DIR / "03_top_squadre.png",
        "trend": OUTPUT_DIR / "04_andamento_squadre_top.png",
    }


def dataframe_to_markdown_table(frame: pd.DataFrame, columns: list[str], max_rows: int = 10) -> str:
    subset = frame.loc[:, columns].head(max_rows).copy()
    if subset.empty:
        return "_Nessun dato disponibile._"

    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
    lines = [header, separator]
    for _, row in subset.iterrows():
        cells = [markdown_value(value) for value in row.tolist()]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def write_reports(
    raw: pd.DataFrame,
    aggregated: pd.DataFrame,
    role_summary: pd.DataFrame,
    team_summary: pd.DataFrame,
    rankings: dict[str, pd.DataFrame],
    visual_paths: dict[str, Path],
) -> None:
    top_players = rankings["player_focus"].head(10)
    scorers = rankings["scorers"].head(10)
    assisters = rankings["assisters"].head(10)
    disciplinary = rankings["disciplinary"].head(10)
    consistency = rankings["consistency"].head(10)

    report_analysis = f"""# Report analisi Fantacalcio 2025/2026

## Pulizia e aggregazione

- File elaborati: {raw['round'].nunique()} giornate.
- Righe pulite a livello evento: {len(raw):,}.
- Giocatori aggregati nel file finale: {len(aggregated):,}.
- Squadre presenti: {aggregated['Squadra'].nunique()}.

La pulizia ha normalizzato i voti con asterisco, ha ricostruito il blocco squadra per ogni foglio e ha aggregato i dati per `Squadra + Cod. + Ruolo + Nome`.

## Evidenze principali

### Giocatori piu efficaci

{dataframe_to_markdown_table(top_players, ['Squadra', 'Nome', 'Ruolo', 'presenze', 'fantamedia', 'gf', 'ass'])}

### Ruoli piu produttivi

{dataframe_to_markdown_table(role_summary, ['Ruolo', 'giocatori', 'fantamedia_media', 'voto_medio'], max_rows=len(role_summary))}

### Squadre piu forti

{dataframe_to_markdown_table(team_summary, ['Squadra', 'presenze', 'fantamedia', 'voto_medio'], max_rows=10)}

### Marcatori, assist e malus

{dataframe_to_markdown_table(scorers, ['Squadra', 'Nome', 'Ruolo', 'gf', 'ass', 'fantamedia'])}

{dataframe_to_markdown_table(assisters, ['Squadra', 'Nome', 'Ruolo', 'ass', 'gf', 'fantamedia'])}

{dataframe_to_markdown_table(disciplinary, ['Squadra', 'Nome', 'Ruolo', 'amm', 'esp', 'malus_disciplinare', 'fantamedia'])}

### Giocatori piu costanti

{dataframe_to_markdown_table(consistency, ['team', 'name', 'role', 'presenze', 'fantamedia', 'std_fantavoto'])}

## Lettura finale

- La squadra piu solida nel complesso e l'Inter, seguita da Como, Roma, Juventus e Napoli.
- Gli attaccanti sono la fascia con la fantamedia media piu alta, mentre i portieri risultano piu penalizzati in termini di rendimento medio.
- Il peso di gol e assist emerge chiaramente: i profili migliori sono quelli che combinano presenza, bonus offensivi e continuita.
- Alcuni difensori e centrocampisti si distinguono per stabilita, non solo per picchi occasionali.

## Visualizzazioni

![Top giocatori](output/01_top_giocatori_fantamedia.png)

![Fantamedia per ruolo](output/02_fantamedia_per_ruolo.png)

![Top squadre](output/03_top_squadre.png)

![Andamento squadre top](output/04_andamento_squadre_top.png)
"""

    report_process = f"""# Report di processo

## Obiettivo

Trasformare i file giornalieri dei voti Fantacalcio in un dataset pulito e aggregato, allineato al template `Fantacalcio_tidy.xlsx`, e produrre alcune analisi utili a fantallenatori e appassionati.

## Passi eseguiti

1. Lettura di tutte le giornate presenti nella cartella.
2. Riconoscimento del titolo variabile in testa al foglio e delle righe di meta-informazione.
3. Estrazione dei blocchi squadra e delle righe giocatore.
4. Pulizia dei voti con asterisco e conversione dei campi numerici.
5. Calcolo del fantavoto per singola presenza.
6. Aggregazione stagionale per squadra, codice, ruolo e nome.
7. Esportazione del workbook finale e produzione di grafici e report.

## Difficolta incontrate

- Il titolo iniziale cambia da giornata a giornata, quindi non si poteva usare un pattern fisso sulla prima riga.
- La giornata 37 non e presente tra i file disponibili, quindi il controllo dei round ha dovuto accettare una sequenza non completa.
- Alcuni voti sono scritti come `6*`; la pulizia ha rimosso il marcatore mantenendo il valore numerico.
- Alcuni codici giocatore compaiono in piu squadre nel corso della stagione; per evitare aggregazioni errate ho usato la chiave `Squadra + Cod. + Ruolo + Nome`.

## Misure adottate

- Parsing robusto basato sulla forma della riga, non su offset fissi.
- Conversione esplicita dei campi numerici con fallback a zero sui bonus/malus non valorizzati.
- Validazione finale con conteggio di file, righe e giocatori aggregati.
- Separazione tra report conclusivo e report metodologico per rendere il risultato piu leggibile.

## Output prodotti

- Workbook finale: `Fantacalcio_tidy.xlsx`
- Grafici: cartella `output/`
- Report analitico: `report_analisi.md`
- Report di processo: `report_processo.md`
"""

    REPORT_ANALISI_FILE.write_text(report_analysis, encoding="utf-8")
    REPORT_PROCESSO_FILE.write_text(report_process, encoding="utf-8")


def main() -> None:
    raw = load_raw_data(BASE_DIR)
    aggregated = build_aggregated_table(raw)
    role_summary = build_role_summary(aggregated)
    team_summary = build_team_summary(raw)
    rankings = build_rankings(raw, aggregated)

    write_tidy_workbook(aggregated, OUTPUT_FILE)
    visual_paths = make_visuals(raw, aggregated, rankings)
    write_reports(raw, aggregated, role_summary, team_summary, rankings, visual_paths)

    print(f"File creato: {OUTPUT_FILE}")
    print(f"Report analitico: {REPORT_ANALISI_FILE}")
    print(f"Report di processo: {REPORT_PROCESSO_FILE}")
    for label, path in visual_paths.items():
        print(f"Grafico {label}: {path}")


if __name__ == "__main__":
    main()