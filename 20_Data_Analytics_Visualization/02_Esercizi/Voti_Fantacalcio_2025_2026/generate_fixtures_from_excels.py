from pathlib import Path
from openpyxl import load_workbook
from analizza_fantacalcio import BASE_DIR, SOURCE_GLOB, is_team_row, is_title_or_meta_row, parse_round_number

OUT_FILE = BASE_DIR / "fixtures.csv"


def extract_team_order(path: Path) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    teams = []
    for row in ws.iter_rows(values_only=True):
        if is_title_or_meta_row(row):
            continue
        if is_team_row(row):
            name = str(row[0]).strip()
            if teams and teams[-1] == name:
                continue
            teams.append(name)
    return teams


def main():
    files = sorted(BASE_DIR.glob(SOURCE_GLOB), key=parse_round_number)
    if not files:
        print("Nessun file giornata trovato.")
        return

    lines = ["round,home,away"]
    for path in files:
        round_no = parse_round_number(path)
        teams = extract_team_order(path)
        if not teams:
            print(f"Attenzione: nessuna squadra trovata in {path.name}")
            lines.append(f"{round_no},,")
            continue
        if len(teams) % 2 != 0:
            print(f"Attenzione: numero dispari squadre in {path.name} ({len(teams)}). L'ultima verra' saltata.")
        pairs = []
        for i in range(0, len(teams) - (len(teams) % 2), 2):
            home = teams[i]
            away = teams[i + 1]
            pairs.append((home, away))

        # For leagues with multiple matches per round, write multiple lines with same round
        for home, away in pairs:
            lines.append(f"{round_no},{home},{away}")

    OUT_FILE.write_text("\n".join(lines), encoding="utf-8")
    print(f"fixtures.csv generato in: {OUT_FILE}")


if __name__ == '__main__':
    main()
